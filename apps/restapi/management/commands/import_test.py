from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
from django.core.management.base import BaseCommand
from apps.users import models as user_models
from apps.companies import models as company_models
from apps.products import models as product_models
from apps.orders import models as order_models
from django.forms import model_to_dict
from django.contrib.gis.geos import Point
# TODO: Импорт только для теста при импорте в продакшене нужно изменить некоторые места


class Command(BaseCommand):

    def handle(self, *args, **options):
        self.delete_objects()
        self.company()
        self.cafe()

    @staticmethod
    def delete_objects():
        companies = company_models.Company.objects.all()
        for company in companies:
            company.delete()
        users = user_models.User.objects.all()
        for user in users:
            user.delete()
        super_user = user_models.User.objects.create_superuser(phone='7777', password='7777')

    @staticmethod
    def company():
        phone = '111111'
        password = '7777'
        company_owner = user_models.User.objects.create_user(phone, password)
        company_owner.user_type = user_models.User.COMPANY_OWNER
        company_owner.save(update_fields=['user_type', ])

        wb2 = load_workbook(filename='export/cafe general settings.xlsx')
        ws2 = wb2['Tablib Dataset']

        company = company_models.Company()
        company.id = 1
        company.owner = company_owner
        company.is_activate = True
        company.name = ws2.cell(row=14, column=4).value
        company.cashback_percent = ws2.cell(row=14, column=13).value
        company.loading_app_image = ws2.cell(row=14, column=11).value
        company.app_image_morning = ws2.cell(row=14, column=11).value
        company.app_image_day = ws2.cell(row=14, column=11).value
        company.app_image_evening = ws2.cell(row=14, column=11).value
        company.save(force_insert=True)

    @staticmethod
    def cafe():
        wb = load_workbook(filename='export/cafe.xlsx')
        ws = wb['Tablib Dataset']
        company = company_models.Company.objects.get(pk=1)
        for r in range(2, ws.max_row + 1):
            if int(ws.cell(row=r, column=26).value) == 1:
                menu_name = ws.cell(row=r, column=4).value + ' menu'
                menu = company_models.Menu.objects.create(
                    name=menu_name,
                    company=company
                )
                cafe = company_models.Cafe.objects.create(
                    company=company_models.Company.objects.get(pk=1),
                    menu=menu,
                    id=ws.cell(row=r, column=1).value,
                    logo=ws.cell(row=r, column=2).value,
                    name=ws.cell(row=r, column=4).value,
                    description=ws.cell(row=r, column=6).value,
                    call_center=ws.cell(row=r, column=7).value,
                    website=ws.cell(row=r, column=8).value,
                    status=ws.cell(row=r, column=9).value,
                    # location=ws.cell(row=r, column=10).value,
                    address=ws.cell(row=r, column=11).value,
                    second_address=ws.cell(row=r, column=12).value,
                    city=ws.cell(row=r, column=13).value,
                    state=ws.cell(row=r, column=14).value,
                    postal_code=ws.cell(row=r, column=15).value,
                    tax_rate=ws.cell(row=r, column=16).value,
                    cafe_timezone=ws.cell(row=r, column=17).value,
                    delivery_available=ws.cell(row=r, column=19).value,
                    delivery_max_distance=ws.cell(row=r, column=20).value,
                    delivery_min_amount=ws.cell(row=r, column=21).value,
                    delivery_fee=ws.cell(row=r, column=22).value,
                    delivery_percent=ws.cell(row=r, column=23).value,
                    delivery_km_amount=ws.cell(row=r, column=24).value,
                    delivery_min_time=ws.cell(row=r, column=25).value,
                    version=ws.cell(row=r, column=27).value,
                    order_limit=ws.cell(row=r, column=28).value,
                    order_time_limit=ws.cell(row=r, column=29).value
                    )
                # create cafe work time
                wb1 = load_workbook(filename='export/week time.xlsx')
                ws1 = wb1['Tablib Dataset']
                for n in range(2, ws1.max_row + 1):
                    if int(ws1.cell(row=n, column=7).value) == cafe.id:
                        week_day = company_models.CafeWeekTime.objects.create(
                            cafe=cafe,
                            day=ws1.cell(row=n, column=2).value,
                            is_open=ws1.cell(row=n, column=3).value,
                            opening_time=ws1.cell(row=n, column=4).value,
                            closing_time=ws1.cell(row=n, column=5).value,
                        )

        cafe1 = company_models.Cafe.objects.get(pk=10)
        cafe1.location = Point(x=float(40.71536), y=float(-74.01100))
        cafe1.save()
        cafe2 = company_models.Cafe.objects.get(pk=13)
        cafe2.location = Point(x=float(40.75506), y=float(-73.98009))
        cafe2.save()
        cafe3 = company_models.Cafe.objects.get(pk=17)
        cafe3.location = Point(x=float(39.75420), y=float(-105.00089))
        cafe3.save()
        cafe4 = company_models.Cafe.objects.get(pk=18)
        cafe4.location = Point(x=float(39.74710), y=float(-104.99881))
        cafe4.save()
        cafe5 = company_models.Cafe.objects.get(pk=22)
        cafe5.location = Point(x=float(40.72110), y=float(-74.00983))
        cafe5.save()

    # @staticmethod
    # def modifier():
    #     wb = load_workbook(filename='export/"modifier category".xlsx')
    #     ws = wb['Tablib Dataset']
    #     for r in range(2, ws.max_row + 1):
    #         if ws.cell(row=r, column=9).value and int(ws.cell(row=r, column=9).value) == 1:
    #             product_models.Modifier.objects.create(
    #                 menu=company_models.Menu.objects.get(pk=1),
    #                 id=ws.cell(row=r, column=1).value,
    #                 name=ws.cell(row=r, column=2).value,
    #                 is_single=ws.cell(row=r, column=5).value,
    #                 required=ws.cell(row=r, column=6).value,
    #                 available=ws.cell(row=r, column=8).value,
    #                 position=1
    #             )
    #
    # @staticmethod
    # def modifier_item():
    #     wb = load_workbook(filename='export/modifier.xlsx')
    #     ws = wb['Tablib Dataset']
    #     mod_list = product_models.Modifier.objects.all().values_list('id', flat=True)
    #     for r in range(2, ws.max_row + 1):
    #         if ws.cell(row=r, column=7).value and int(ws.cell(row=r, column=7).value) in mod_list:
    #             product_models.ModifierItem.objects.create(
    #                 modifier=product_models.Modifier.objects.get(pk=int(ws.cell(row=r, column=7).value)),
    #                 id=ws.cell(row=r, column=1).value,
    #                 name=ws.cell(row=r, column=2).value,
    #                 price=ws.cell(row=r, column=3).value,
    #                 available=ws.cell(row=r, column=4).value,
    #                 default=ws.cell(row=r, column=8).value,
    #                 position=ws.cell(row=r, column=10).value,
    #             )
    #
    @staticmethod
    def product_category():
        company = company_models.Company.objects.all()
        all_product_menu = product_models.Menu.objects.create(
            name='all meals',
            company=company
        )

        wbcat = load_workbook(filename='export/product category.xlsx')
        wscat = wbcat['Tablib Dataset']
        for r in range(2, wscat.max_row + 1):
            if wscat.cell(row=r, column=10).value and int(wscat.cell(row=r, column=10).value) == 1:
                product_models.ProductCategory.objects.create(
                    menu=all_product_menu,
                    id=wscat.cell(row=r, column=1).value,
                    name=wscat.cell(row=r, column=3).value,
                    image=wscat.cell(row=r, column=6).value,
                    available=wscat.cell(row=r, column=7).value,
                    start=wscat.cell(row=r, column=11).value,
                    end=wscat.cell(row=r, column=12).value,
                    position=wscat.cell(row=r, column=8).value,
                    is_kitchen=True
                )

    @staticmethod
    def product():
        wb = load_workbook(filename='export/product.xlsx')
        ws = wb['Tablib Dataset']
        cat_list = product_models.ProductCategory.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=9).value and int(ws.cell(row=r, column=9).value) in cat_list:
                product_models.Product.objects.create(
                        id=ws.cell(row=r, column=1).value,
                        category=product_models.ProductCategory.objects.get(
                            pk=int(ws.cell(row=r, column=9).value)),
                        image=ws.cell(row=r, column=2).value,
                        name=ws.cell(row=r, column=3).value,
                        description=ws.cell(row=r, column=4).value,
                        position=ws.cell(row=r, column=14).value
                    )
    #
    # @staticmethod
    # def product_size():
    #     wb = load_workbook(filename='export/size.xlsx')
    #     ws = wb['Tablib Dataset']
    #     prod_list = product_models.Product.objects.all().values_list('id', flat=True)
    #     for r in range(2, ws.max_row + 1):
    #         if ws.cell(row=r, column=4).value and int(ws.cell(row=r, column=4).value) in prod_list:
    #             product_models.ProductSize.objects.create(
    #                 id=ws.cell(row=r, column=1).value,
    #                 product=product_models.Product.objects.get(
    #                     pk=int(ws.cell(row=r, column=4).value)),
    #                 name=ws.cell(row=r, column=2).value,
    #                 price=ws.cell(row=r, column=3).value,
    #                 available=ws.cell(row=r, column=6).value,
    #                 default=ws.cell(row=r, column=5).value,
    #
    #             )
    #
    # @staticmethod
    # def product_add_modifier():
    #     wb = load_workbook(filename='export/product modifier.xlsx')
    #     ws = wb['Tablib Dataset']
    #     prod_list = product_models.Product.objects.all().values_list('id', flat=True)
    #     mod_list = product_models.Modifier.objects.all().values_list('id', flat=True)
    #     for r in range(2, ws.max_row + 1):
    #
    #         a = ws.cell(row=r, column=3).value and int(ws.cell(row=r, column=3).value) in prod_list
    #         b = ws.cell(row=r, column=2).value and int(ws.cell(row=r, column=2).value) in mod_list
    #         if a and b:
    #             product = product_models.Product.objects.get(pk=int(ws.cell(row=r, column=3).value))
    #             modifier = product_models.Modifier.objects.get(pk=int(ws.cell(row=r, column=2).value))
    #             product.modifiers.add(modifier)
    #
    # @staticmethod
    # def product_add_time():
    #     wb = load_workbook(filename='export/cafe meals.xlsx')
    #     ws = wb['Tablib Dataset']
    #     prod_list = product_models.Product.objects.all().values_list('id', flat=True)
    #
    #     for r in range(2, ws.max_row + 1):
    #
    #         if ws.cell(row=r, column=3).value and int(ws.cell(row=r, column=3).value) in prod_list:
    #             product = product_models.Product.objects.get(pk=int(ws.cell(row=r, column=3).value))
    #             if ws.cell(row=r, column=4).value:
    #                 product.start = ws.cell(row=r, column=4).value
    #             if ws.cell(row=r, column=5).value:
    #                 product.end = ws.cell(row=r, column=5).value
    #             if ws.cell(row=r, column=6).value:
    #                 product.quickest_time = ws.cell(row=r, column=6).value
    #             if ws.cell(row=r, column=7).value:
    #                 product.tax_present = ws.cell(row=r, column=7).value
    #
    #             product.save()



