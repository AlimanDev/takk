from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
from django.core.management.base import BaseCommand
from apps.users import models as user_models
from apps.companies import models as company_models
from apps.products import models as product_models
from apps.orders import models as order_models
from django.forms import model_to_dict
from django.contrib.gis.geos import Point
# TODO: Импорт только для теста при импорте в продакшене нужно изменить некоторые места


class Command(BaseCommand):

    def handle(self, *args, **options):
        self.delete_objects()
        # self.user()
        self.company()
        self.menu()
        self.cafe()
        self.cafe_week_time()
        self.modifier()
        self.modifier_item()
        self.product_category()
        self.product()
        self.product_size()
        # self.favorite_cart()
        # self.favorite_cart_item()
        # self.order()
        # self.order_items()
        self.create_menu_duplicate()
        self.update_location()

    @staticmethod
    def delete_objects():
        orders = order_models.Order.objects.all()
        for order in orders:
            order.delete()
        users = user_models.User.objects.all().exclude(phone='7777777')
        for user in users:
            user.delete()
        company = company_models.Company.objects.all()
        for item in company:
            item.delete()
        super_user = user_models.User.objects.create_superuser(phone='7777', password='7777')

    @staticmethod
    def user():
        wb = load_workbook(filename='export/user.xlsx')
        ws = wb['Tablib Dataset']
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=12).value:
                user_models.User.objects.create(
                    id=ws.cell(row=r, column=1).value,
                    # password=ws.cell(row=r, column=2).value,
                    # email=ws.cell(row=r, column=7).value,  # TODO: Не работает исправить
                    username='{} {}'.format(ws.cell(row=r, column=14).value, ws.cell(row=r, column=15).value),
                    stripe_card_id=ws.cell(row=r, column=11).value,
                    phone=ws.cell(row=r, column=12).value,
                    phone_is_verified=True,
                    date_of_birthday=ws.cell(row=r, column=18).value,
                    avatar=ws.cell(row=r, column=20).value,
                    referral_code=ws.cell(row=r, column=21).value,
                    batch_payment=ws.cell(row=r, column=23).value,
                    # date_joined=ws.cell(row=r, column=16).value

                )

    @staticmethod
    def company():
        wb = load_workbook(filename='export/company.xlsx')
        ws = wb['Tablib Dataset']
        wb2 = load_workbook(filename='export/cafe general settings.xlsx')
        ws2 = wb2['Tablib Dataset']
        for r in range(2, 3):
            company_models.Company.objects.create(
                id=ws.cell(row=r, column=1).value,
                owner=user_models.User.objects.get(phone='7777'),
                # owner=user_models.User.objects.get(pk=1),
                is_activate=True,
                name=ws2.cell(row=14, column=4).value,
                cashback_percent=ws2.cell(row=14, column=13).value,
                loading_app_image=ws2.cell(row=14, column=11).value,
                app_image_morning=ws2.cell(row=14, column=11).value,
                app_image_day=ws2.cell(row=14, column=11).value,
                app_image_evening=ws2.cell(row=14, column=11).value
            )

    @staticmethod
    def menu():
        company_models.Menu.objects.create(
            id=1,
            company=company_models.Company.objects.get(pk=1),
            name='menu 1'
        )

    @staticmethod
    def cafe():
        wb = load_workbook(filename='export/cafe.xlsx')
        ws = wb['Tablib Dataset']
        for r in range(2, ws.max_row + 1):
            if int(ws.cell(row=r, column=26).value) == 1:
                company_models.Cafe.objects.create(
                    company=company_models.Company.objects.get(pk=1),
                    # menu=company_models.Menu.objects.get(pk=1),
                    id=ws.cell(row=r, column=1).value,
                    logo=ws.cell(row=r, column=2).value,
                    name=ws.cell(row=r, column=4).value,
                    description=ws.cell(row=r, column=6).value,
                    call_center=ws.cell(row=r, column=7).value,
                    website=ws.cell(row=r, column=8).value,
                    status=ws.cell(row=r, column=9).value,
                    # location=ws.cell(row=r, column=10).value,
                    address=ws.cell(row=r, column=11).value,
                    second_address=ws.cell(row=r, column=12).value,
                    city=ws.cell(row=r, column=13).value,
                    state=ws.cell(row=r, column=14).value,
                    postal_code=ws.cell(row=r, column=15).value,
                    tax_rate=ws.cell(row=r, column=16).value,
                    cafe_timezone=ws.cell(row=r, column=17).value,
                    delivery_available=ws.cell(row=r, column=19).value,
                    delivery_max_distance=ws.cell(row=r, column=20).value,
                    delivery_min_amount=ws.cell(row=r, column=21).value,
                    delivery_fee=ws.cell(row=r, column=22).value,
                    delivery_percent=ws.cell(row=r, column=23).value,
                    delivery_km_amount=ws.cell(row=r, column=24).value,
                    delivery_min_time=ws.cell(row=r, column=25).value,
                    version=ws.cell(row=r, column=27).value,
                    order_limit=ws.cell(row=r, column=28).value,
                    order_time_limit=ws.cell(row=r, column=29).value
                    )

    @staticmethod
    def cafe_week_time():
        wb = load_workbook(filename='export/week time.xlsx')
        ws = wb['Tablib Dataset']
        cafe_list = company_models.Cafe.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):
            if int(ws.cell(row=r, column=7).value) in cafe_list:

                company_models.CafeWeekTime.objects.create(
                    cafe=company_models.Cafe.objects.get(pk=int(ws.cell(row=r, column=7).value)),
                    # id=ws.cell(row=r, column=1).value,
                    day=ws.cell(row=r, column=2).value,
                    is_open=ws.cell(row=r, column=3).value,
                    opening_time=ws.cell(row=r, column=4).value,
                    closing_time=ws.cell(row=r, column=5).value,
                )

    @staticmethod
    def modifier():
        wb = load_workbook(filename='export/modifier category.xlsx')
        ws = wb['Tablib Dataset']
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=9).value and int(ws.cell(row=r, column=9).value) == 1:
                product_models.Modifier.objects.create(
                    menu=company_models.Menu.objects.get(pk=1),
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value,
                    is_single=ws.cell(row=r, column=5).value,
                    required=ws.cell(row=r, column=6).value,
                    available=ws.cell(row=r, column=8).value,
                    position=1
                )

    @staticmethod
    def modifier_item():
        wb = load_workbook(filename='export/modifier.xlsx')
        ws = wb['Tablib Dataset']
        mod_list = product_models.Modifier.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=7).value and int(ws.cell(row=r, column=7).value) in mod_list:
                product_models.ModifierItem.objects.create(
                    modifier=product_models.Modifier.objects.get(pk=int(ws.cell(row=r, column=7).value)),
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value,
                    price=ws.cell(row=r, column=3).value,
                    available=ws.cell(row=r, column=4).value,
                    default=ws.cell(row=r, column=8).value,
                    position=ws.cell(row=r, column=10).value,
                )

    @staticmethod
    def product_category():
        wb = load_workbook(filename='export/product category.xlsx')
        ws = wb['Tablib Dataset']
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=10).value and int(ws.cell(row=r, column=10).value) == 1:
                product_models.ProductCategory.objects.create(
                    menu=company_models.Menu.objects.get(pk=1),
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=3).value,
                    image=ws.cell(row=r, column=6).value,
                    available=ws.cell(row=r, column=7).value,
                    start=ws.cell(row=r, column=11).value,
                    end=ws.cell(row=r, column=12).value,
                    position=ws.cell(row=r, column=8).value,
                    is_kitchen=True
                )

    @staticmethod
    def product():
        wb = load_workbook(filename='export/product.xlsx')
        ws = wb['Tablib Dataset']
        cat_list = product_models.ProductCategory.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=9).value and int(ws.cell(row=r, column=9).value) in cat_list:
                product_models.Product.objects.create(
                        id=ws.cell(row=r, column=1).value,
                        category=product_models.ProductCategory.objects.get(
                            pk=int(ws.cell(row=r, column=9).value)),
                        image=ws.cell(row=r, column=2).value,
                        name=ws.cell(row=r, column=3).value,
                        description=ws.cell(row=r, column=4).value,
                        position=ws.cell(row=r, column=14).value
                    )

    @staticmethod
    def product_size():
        wb = load_workbook(filename='export/size.xlsx')
        ws = wb['Tablib Dataset']
        prod_list = product_models.Product.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=4).value and int(ws.cell(row=r, column=4).value) in prod_list:
                product_models.ProductSize.objects.create(
                    id=ws.cell(row=r, column=1).value,
                    product=product_models.Product.objects.get(
                        pk=int(ws.cell(row=r, column=4).value)),
                    name=ws.cell(row=r, column=2).value,
                    price=ws.cell(row=r, column=3).value,
                    available=ws.cell(row=r, column=6).value,
                    default=ws.cell(row=r, column=5).value,

                )

    @staticmethod
    def product_add_modifier():
        wb = load_workbook(filename='export/product modifier.xlsx')
        ws = wb['Tablib Dataset']
        prod_list = product_models.Product.objects.all().values_list('id', flat=True)
        mod_list = product_models.Modifier.objects.all().values_list('id', flat=True)
        for r in range(2, ws.max_row + 1):

            a = ws.cell(row=r, column=3).value and int(ws.cell(row=r, column=3).value) in prod_list
            b = ws.cell(row=r, column=2).value and int(ws.cell(row=r, column=2).value) in mod_list
            if a and b:
                product = product_models.Product.objects.get(pk=int(ws.cell(row=r, column=3).value))
                modifier = product_models.Modifier.objects.get(pk=int(ws.cell(row=r, column=2).value))
                product.modifiers.add(modifier)

    @staticmethod
    def product_add_time():
        wb = load_workbook(filename='export/cafe meals.xlsx')
        ws = wb['Tablib Dataset']
        prod_list = product_models.Product.objects.all().values_list('id', flat=True)

        for r in range(2, ws.max_row + 1):

            if ws.cell(row=r, column=3).value and int(ws.cell(row=r, column=3).value) in prod_list:
                product = product_models.Product.objects.get(pk=int(ws.cell(row=r, column=3).value))
                if ws.cell(row=r, column=4).value:
                    product.start = ws.cell(row=r, column=4).value
                if ws.cell(row=r, column=5).value:
                    product.end = ws.cell(row=r, column=5).value
                if ws.cell(row=r, column=6).value:
                    product.quickest_time = ws.cell(row=r, column=6).value
                if ws.cell(row=r, column=7).value:
                    product.tax_present = ws.cell(row=r, column=7).value

                product.save()

    @staticmethod
    def favorite_cart():
        wb = load_workbook(filename='export/favorite cart.xlsx')
        ws = wb['Tablib Dataset']
        cafe_list = company_models.Cafe.objects.all().values_list('id', flat=True)

        for r in range(2, ws.max_row + 1):

            if ws.cell(row=r, column=4).value and int(ws.cell(row=r, column=4).value) in cafe_list:
                cafe = company_models.Cafe.objects.get(pk=int(ws.cell(row=r, column=4).value))
                user = user_models.User.objects.get(pk=int(ws.cell(row=r, column=2).value))
                instance = order_models.FavoriteCart()
                instance.cafe = cafe
                instance.user = user
                instance.name = ws.cell(row=r, column=3).value
                if ws.cell(row=r, column=5).value:
                    delivery = order_models.Delivery.objects.get(pk=int(ws.cell(row=r, column=5).value))
                    instance.delivery = delivery
                instance.save()

    @staticmethod
    def favorite_cart_item():
        wb = load_workbook(filename='export/favorite cart item.xlsx')
        ws = wb['Tablib Dataset']
        favorite_list = order_models.FavoriteCart.objects.all().values_list('id', flat=True)

        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=2).value and int(ws.cell(row=r, column=2).value) in favorite_list:
                favorite_cart = order_models.FavoriteCart.objects.get(
                    pk=int(ws.cell(row=r, column=2).value))
                size = product_models.ProductSize.objects.get(
                    pk=int(ws.cell(row=r, column=3).value))
                instance = order_models.FavoriteCartItem()
                instance.id = ws.cell(row=r, column=1).value
                instance.favorite_cart = favorite_cart
                instance.product = size.product
                instance.size = size
                instance.quantity = int(ws.cell(row=r, column=5).value)
                instance.instruction = ws.cell(row=6, column=6).value
                instance.save()
                modifier = list(ws.cell(row=r, column=4).value)
                for item in modifier:
                    mod = product_models.ModifierItem.objects.get(pk=item)
                    instance.product_modifiers.add(mod)

    @staticmethod
    def order():
        wb = load_workbook(filename='export/order.xlsx')
        ws = wb['Tablib Dataset']
        cafe_list = company_models.Cafe.objects.all().values_list('id', flat=True)

        for r in range(2, ws.max_row + 1):
            a = ws.cell(row=r, column=9).value and int(ws.cell(row=r, column=9).value) in cafe_list
            b = ws.cell(row=r, column=2).value
            if a and b:

                cafe = company_models.Cafe.objects.get(
                    pk=int(ws.cell(row=r, column=9).value))
                user = user_models.User.objects.get(
                    pk=int(ws.cell(row=r, column=2).value))
                order = order_models.Order.objects.create(
                    id=ws.cell(row=r, column=1).value,
                    customer=user,
                    cafe=cafe,

                    order_unique_id=ws.cell(row=r, column=4).value,
                    sub_total_price=ws.cell(row=r, column=5).value,
                    tax_total=ws.cell(row=r, column=6).value,
                    total_price=ws.cell(row=r, column=7).value,
                    status=ws.cell(row=r, column=8).value,
                    pre_order=ws.cell(row=r, column=10).value,
                    pre_order_date=ws.cell(row=r, column=11).value,
                    tip=ws.cell(row=r, column=14).value,
                    tip_percent=ws.cell(row=r, column=15).value,
                    is_confirm=ws.cell(row=r, column=18).value,
                    created=ws.cell(row=r, column=3).value,
                )
                # if ws.cell(row=r, column=16).value:
                #     delivery = order_models.Delivery.objects.get(
                #         pk=ws.cell(row=r, column=17).value)
                #     order.delivery = delivery
                #     order.save()

    @staticmethod
    def order_items():
        wb = load_workbook(filename='export/cart.xlsx')
        ws = wb['Tablib Dataset']
        order_list = order_models.Order.objects.all().values_list('id', flat=True)

        for r in range(2, ws.max_row + 1):
            a = ws.cell(row=r, column=5).value and int(ws.cell(row=r, column=5).value) in order_list
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            if a and b and c:
                order = order_models.Order.objects.get(pk=int(ws.cell(row=r, column=5).value))
                product = product_models.Product.objects.get(pk=int(ws.cell(row=r, column=2).value))
                product_size = product_models.ProductSize.objects.get(pk=int(ws.cell(row=r, column=3).value))
                instance = order_models.OrderItem.objects.create(
                    id=ws.cell(row=r, column=1).value,
                    order=order,
                    product=product,
                    product_size=product_size,
                    quantity=int(ws.cell(row=r, column=4).value),
                    is_free=ws.cell(row=r, column=6).value,
                    product_price=product_size.price,
                    product_name=product.name,

                    tax_percent=ws.cell(row=r, column=10).value,
                    free_count=ws.cell(row=r, column=9).value,
                    tax_rate=ws.cell(row=r, column=9).value
                )

                if ws.cell(row=r, column=8).value:
                    instance.instruction = ws.cell(row=r, column=8).value
                    instance.save()

    @staticmethod
    def create_menu_duplicate():


        for i in range(0, 4):
            company_models.Menu.objects.create(
                id=i+2,
                name='menu {}'.format(i+2),
                company=company_models.Company.objects.get(pk=1)
            )
        menus = company_models.Menu.objects.all().exclude(pk=1)
        menu = company_models.Menu.objects.get(pk=1)

        for new_menu in menus:

            for prod_cat in menu.categories.all():
                new_cat = product_models.ProductCategory.objects.get(id=prod_cat.id)
                last_id = product_models.ProductCategory.objects.latest('id').pk + 1
                new_cat.id = None
                new_cat.id = last_id
                new_cat.menu = new_menu
                new_cat.save(force_insert=True)

                for prod in prod_cat.products.all():
                    new_prod = product_models.Product.objects.get(id=prod.id)
                    last_id = product_models.Product.objects.latest('id').pk + 1
                    new_prod.id = last_id
                    new_prod.category = new_cat
                    new_prod.save(force_insert=True)


                    for size in prod.sizes.all():
                        new_size = product_models.ProductSize.objects.get(id=size.id)
                        last_id = product_models.ProductSize.objects.latest('id').pk + 1
                        new_size.id = last_id
                        new_size.product = new_prod
                        new_size.save(force_insert=True)

                    for prod_mod in prod.modifiers.all():
                        new_mod = product_models.Modifier.objects.get(id=prod_mod.id)
                        last_id = product_models.Modifier.objects.latest('id').pk + 1
                        new_mod.id = last_id
                        new_mod.menu = new_cat.menu
                        new_mod.save(force_insert=True)
                        new_prod.modifiers.add(new_mod)


                        for mod_item in prod_mod.items.all():
                            new_mod_item = product_models.ModifierItem.objects.get(id=mod_item.id)
                            last_id = product_models.ModifierItem.objects.latest('id').pk + 1
                            new_mod_item.id = last_id
                            new_mod_item.modifier = new_mod
                            new_mod_item.save(force_insert=True)

        cafe1 = company_models.Cafe.objects.get(pk=22)
        menu1 = company_models.Menu.objects.get(pk=1)
        cafe1.menu = menu1
        cafe1.save()
        cafe2 = company_models.Cafe.objects.get(pk=18)
        menu2 = company_models.Menu.objects.get(pk=2)
        cafe2.menu = menu2
        cafe2.save()
        cafe3 = company_models.Cafe.objects.get(pk=17)
        menu3 = company_models.Menu.objects.get(pk=3)
        cafe3.menu = menu3
        cafe3.save()
        cafe4 = company_models.Cafe.objects.get(pk=13)
        menu4 = company_models.Menu.objects.get(pk=4)
        cafe4.menu = menu4
        cafe4.save()
        cafe5 = company_models.Cafe.objects.get(pk=10)
        menu5 = company_models.Menu.objects.get(pk=5)
        cafe5.menu = menu5
        cafe5.save()

    @staticmethod
    def update_location():
        cafe1 = company_models.Cafe.objects.get(pk=10)
        cafe1.location = Point(x=float(40.71536), y=float(-74.01100))
        cafe1.save()
        cafe2 = company_models.Cafe.objects.get(pk=13)
        cafe2.location = Point(x=float(40.75506), y=float(-73.98009))
        cafe2.save()
        cafe3 = company_models.Cafe.objects.get(pk=17)
        cafe3.location = Point(x=float(39.75420), y=float(-105.00089))
        cafe3.save()
        cafe4 = company_models.Cafe.objects.get(pk=18)
        cafe4.location = Point(x=float(39.74710), y=float(-104.99881))
        cafe4.save()
        cafe5 = company_models.Cafe.objects.get(pk=22)
        cafe5.location = Point(x=float(40.72110), y=float(-74.00983))
        cafe5.save()